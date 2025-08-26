import pytest
from playwright.sync_api import sync_playwright, Page
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image  # เพิ่มการนำเข้า Image สำหรับจัดการรูปภาพใน Excel
import os
import time

# สร้างหรือโหลดไฟล์ Excel สำหรับบันทึกผลการทดสอบ 
def setup_excel():
    filename = "TestSheet.xlsx"
    wb = load_workbook(filename)
    if "Test Results" not in wb.sheetnames: # ตรวจสอบว่ามีชีท "Test Results" หรือไม่
        ws = wb.create_sheet("Test Results") # สร้างชีทใหม่ชื่อ "Test Results"
        ws.append(["Test Case", "Status", "Screenshot"]) # เพิ่มหัวคอลัมน์
        wb.save(filename) # บันทึกไฟล์ Excel
    return filename

# บันทึกผลลัพธ์ลง Excel โดยค้นหาชื่อเทสต์เคสในคอลัมน์ A 
def save_result_custom_row(test_name, status, screenshot_path):
    filename = setup_excel() # เรียกใช้ฟังก์ชัน setup_excel เพื่อเตรียมไฟล์ Excel
    wb = load_workbook(filename) # โหลดไฟล์ Excel
    ws = wb["Test Results"] # เลือกชีท "Test Results"

    # แยกชื่อเทสต์เคสจาก test_name (เช่น test_login01_pos01 -> pos01)
    test_case_id = test_name.split("_")[-1]

    # ค้นหาแถวที่ตรงกับชื่อเทสต์เคสในคอลัมน์ A
    target_row = None
    for row in ws.iter_rows(min_row=1, max_col=1):
        if row[0].value == test_case_id:  # ตรวจสอบค่าในคอลัมน์ A
            target_row = row[0].row  # เก็บหมายเลขแถว
            break

    # บันทึกข้อมูลที่ column H (status) และ column I (screenshot) ในแถวที่พบ
    ws[f"H{target_row}"] = status

    # แสดงข้อมูลในคอนโซลเพื่อการดีบั๊ก
    print(f"Status: {status}")
    print(f"Screenshot Path: {screenshot_path}")

    # แทรกรูปภาพใน Excel
    if os.path.exists(screenshot_path):
        img = Image(screenshot_path)
        img.width = 250
        img.height = 190
        ws.add_image(img, f"I{target_row}")

    # บันทึกไฟล์ Excel
    wb.save(filename)

# จับภาพหน้าจอและบันทึกไฟล์
def capture_screenshot(page, test_name):
    screenshot_path = f"screenshots/{test_name}.png"
    
    # ตรวจสอบว่าโฟลเดอร์ screenshots/ มีอยู่หรือไม่ ถ้าไม่มีให้สร้าง
    os.makedirs("screenshots", exist_ok=True)
    
    try:
        # รอให้หน้าโหลดหรือองค์ประกอบที่ต้องการปรากฏก่อนจับภาพ
        page.wait_for_selector("input[type='submit'][value=' เข้าสู่ระบบ ']", timeout=10000)  # Adjust as necessary
        page.screenshot(path=screenshot_path)
        print(f"Screenshot saved to {screenshot_path}")  # ตรวจสอบว่าไฟล์ถูกบันทึกหรือไม่
    except Exception as e:
        print(f"Error capturing screenshot: {e}")
        screenshot_path = "Screenshot Failed"
    
    return screenshot_path

# สร้างเบราว์เซอร์ Chromium สำหรับการทดสอบ
@pytest.fixture(scope="function")
def browser():
    with sync_playwright() as p:
        browser = p.chromium.launch(headless=False) # เปิดเบราว์เซอร์แบบเห็น GUI (headless=False)
        yield browser
        browser.close() # ปิดเบราว์เซอร์หลังการทดสอบ

# สร้างหน้าเว็บใหม่ในเบราว์เซอร์
@pytest.fixture(scope="function")
def page(browser):
    context = browser.new_context()
    page = context.new_page()
    yield page
    context.close()


def test_login01_pos01(page: Page):
    test_name = "test_login01_pos01"
    try:
        page.goto("https://reg.rmutk.ac.th/registrar/login.asp")

        # รอให้มีข้อความ "เปลี่ยนรหัสผ่าน" ปรากฏ
        success_message = page.wait_for_selector(':has-text("เปลี่ยนรหัสผ่าน")', timeout=60000)
        assert success_message.is_visible()

        # บันทึกผลลัพธ์การทดสอบลงในไฟล์ Excel โดยค้นหาแถวจากชื่อเทสต์เคสในคอลัมน์
        screenshot = capture_screenshot(page, test_name)
        save_result_custom_row(test_name, "Pass", screenshot)
    except Exception:
        screenshot = capture_screenshot(page, test_name)
        save_result_custom_row(test_name, "Fail", screenshot)
        raise

def test_login01_neg01(page: Page):
    test_name = "test_login01_neg01"
    try:
        page.goto("https://reg.rmutk.ac.th/registrar/login.asp")
        page.locator('input[name="f_uid"]').fill("User_id") #ใส่รหัสนักศึกษาถูก
        page.locator('input[name="f_pwd"]').fill("5555") #ใส่รหัสผ่านผิด
        page.locator('input[name="f_idcard"]').fill("5555")#ใส่เลขบัตรประชาชนผิด
        page.locator('input[type="submit"][value=" เข้าสู่ระบบ "]').click()

        error_message = page.wait_for_selector(
            "text=ก รุ ณ า ป้ อ น ร หั ส ป ร ะ จ ำ ตั ว น ศ .ร หั ส บั ต ร ป ร ะ ช า ช น แ ล ะ ร หั ส ผ่ า น ใ ห้ ถู ก ต้ อ ง",
            timeout=60000
        )
        color = error_message.evaluate('element => window.getComputedStyle(element).color')
        assert error_message.is_visible()
        assert color == "rgb(0,0,255)"

        # บันทึกผลลัพธ์การทดสอบลงในไฟล์ Excel โดยค้นหาแถวจากชื่อเทสต์เคสในคอลัมน์
        screenshot = capture_screenshot(page, test_name)
        save_result_custom_row(test_name, "Pass", screenshot)
    except Exception:
        screenshot = capture_screenshot(page, test_name)
        save_result_custom_row(test_name, "Fail", screenshot)
        raise


def test_login01_neg02(page: Page):
    test_name = "test_login01_neg02"
    try:
        page.goto("https://reg.rmutk.ac.th/registrar/login.asp")
        page.locator('input[name="f_uid"]').fill("") #ไม่กรอก
        page.locator('input[name="f_pwd"]').fill("")#ไม่กรอก
        page.locator('input[name="f_idcard"]').fill("")  #ไม่กรอก
        page.locator('input[type="submit"][value=" เข้าสู่ระบบ "]').click()

        # ตรวจสอบว่ามีข้อความแจ้งเตือน
        error_message = page.wait_for_selector("text=กรุณาป้อนรหัสประจำตัวและรหัสผ่านให้ถูกต้อง",timeout=60000)
        assert error_message.is_visible()

        # บันทึกผลลัพธ์การทดสอบลงในไฟล์ Excel โดยค้นหาแถวจากชื่อเทสต์เคสในคอลัมน์
        screenshot = capture_screenshot(page, test_name)
        save_result_custom_row(test_name, "Pass", screenshot)
    except Exception:
        screenshot = capture_screenshot(page, test_name)
        save_result_custom_row(test_name, "Fail", screenshot)
        raise

def test_login01_neg03(page: Page):
    test_name = "test_login01_neg03"
    try:
        page.goto("https://reg.rmutk.ac.th/registrar/login.asp")
        page.locator('input[name="f_uid"]').fill("User_id")#ใส่รหัสนักศึกษาถูก
        page.locator('input[name="f_pwd"]').fill("รหัสผ่านจริง")# ใส่รหัสผ่านถูก
        page.locator('input[name="f_idcard"]').fill("5555")  # ใส่เลขบัตรประชาชนผิด
        page.locator('input[type="submit"][value=" เข้าสู่ระบบ "]').click()

        error_message = page.wait_for_selector("text=ก รุ ณ า ป้ อ น ร หั ส ป ร ะ จ ำ ตั ว น ศ .ร หั ส บั ต ร ป ร ะ ช า ช น แ ล ะ ร หั ส ผ่ า น ใ ห้ ถู ก ต้ อ ง",timeout=60000)
        assert error_message.is_visible()

        # บันทึกผลลัพธ์การทดสอบลงในไฟล์ Excel โดยค้นหาแถวจากชื่อเทสต์เคสในคอลัมน์
        screenshot = capture_screenshot(page, test_name)
        save_result_custom_row(test_name, "Pass", screenshot)
    except Exception:
        screenshot = capture_screenshot(page, test_name)
        save_result_custom_row(test_name, "Fail", screenshot)
        raise

def test_login01_neg04(page: Page):
    test_name = "test_login01_neg04"
    try:
        page.goto("https://reg.rmutk.ac.th/registrar/login.asp")
        page.locator('input[name="f_uid"]').fill("5555")  # ใส่รหัสนักศึกษาผิด
        page.locator('input[name="f_pwd"]').fill("รหัสผ่านจริง")  # ใส่รหัสผ่านถูก
        page.locator('input[name="f_idcard"]').fill("บัตร ปชช")  # ใส่เลขบัตรประชาชนถูก
        page.locator('input[type="submit"][value=" เข้าสู่ระบบ "]').click()

        error_message = page.wait_for_selector("text=กรุณาป้อนรหัสประจำตัวและรหัสผ่านให้ถูกต้อง",timeout=60000)
        assert error_message.is_visible()

        # บันทึกผลลัพธ์การทดสอบลงในไฟล์ Excel โดยค้นหาแถวจากชื่อเทสต์เคสในคอลัมน์
        screenshot = capture_screenshot(page, test_name)
        save_result_custom_row(test_name, "Pass", screenshot)
    except Exception:
        screenshot = capture_screenshot(page, test_name)
        save_result_custom_row(test_name, "Fail", screenshot)
        raise

def test_login01_neg05(page: Page):
    test_name = "test_login01_neg05"
    try:
        page.goto("https://reg.rmutk.ac.th/registrar/login.asp")
        page.locator('input[name="f_uid"]').fill("5555")  # ใส่รหัสนักศึกษาผิด
        page.locator('input[name="f_pwd"]').fill("รหัสผ่านจริง")  # ใส่รหัสผ่านถูก
        page.locator('input[name="f_idcard"]').fill("บชช")  # ใส่เลขบัตรประชาชนถูก
        page.locator('input[type="submit"][value=" เข้าสู่ระบบ "]').click()

        error_message = page.wait_for_selector("text=กรุณาป้อนรหัสประจำตัวและรหัสผ่านให้ถูกต้อง",timeout=60000)
        assert error_message.is_visible()

        # บันทึกผลลัพธ์การทดสอบลงในไฟล์ Excel โดยค้นหาแถวจากชื่อเทสต์เคสในคอลัมน์
        screenshot = capture_screenshot(page, test_name)
        save_result_custom_row(test_name, "Pass", screenshot)
    except Exception:
        screenshot = capture_screenshot(page, test_name)
        save_result_custom_row(test_name, "Fail", screenshot)
        raise

