import pytest
from playwright.sync_api import sync_playwright, Page
from openpyxl import Workbook, load_workbook
import os
import time

def setup_excel():
    filename = "TestSheet.xlsx"
    wb = load_workbook(filename)
    if "Test Results" not in wb.sheetnames:
        ws = wb.create_sheet("Test Results")
        ws.append(["Test Case", "Status", "Screenshot"])
        wb.save(filename)
    return filename

def save_result(test_name, status, screenshot_path):
    filename = setup_excel()
    wb = load_workbook(filename)
    ws = wb["Test Results"]
    ws.append([test_name, status, screenshot_path])
    wb.save(filename)

def capture_screenshot(page, test_name):
    screenshot_path = f"screenshots/{test_name}.png"
    
    # ตรวจสอบว่าโฟลเดอร์ screenshots/ มีอยู่หรือไม่ ถ้าไม่มีให้สร้าง
    os.makedirs("screenshots", exist_ok=True)
    
    try:
        # รอให้หน้าโหลดหรือองค์ประกอบที่ต้องการปรากฏก่อนจับภาพ
        page.wait_for_selector("input[type='submit'][value=' เข้าสู่ระบบ ']", timeout=10000)  # Adjust as necessary
        page.screenshot(path=screenshot_path)
        print(f"Screenshot saved to {screenshot_path}")  # ตรวจสอบว่าไฟล์ถูกบันทึกหรือไม่
    except Exception as e:
        print(f"Error capturing screenshot: {e}")
        screenshot_path = "Screenshot Failed"
    
    return screenshot_path

@pytest.fixture(scope="function")
def browser():
    with sync_playwright() as p:
        browser = p.chromium.launch(headless=False)
        yield browser
        browser.close()

@pytest.fixture(scope="function")
def page(browser):
    context = browser.new_context()
    page = context.new_page()
    yield page
    context.close()

def test_login01_neg01(page: Page):
    test_name = "test_login01_neg01"
    try:
        page.goto("https://reg.rmutk.ac.th/registrar/login.asp")
        page.locator('input[name="f_uid"]').fill("User_id")
        page.locator('input[name="f_pwd"]').fill("5555")
        page.locator('input[name="f_idcard"]').fill("5555")
        page.locator('input[type="submit"][value=" เข้าสู่ระบบ "]').click()
        error_message = page.wait_for_selector("text=ก รุ ณ า ป้ อ น ร หั ส ป ร ะ จ ำ ตั ว น ศ .ร หั ส บั ต ร ป ร ะ ช า ช น แ ล ะ ร หั ส ผ่ า น ใ ห้ ถู ก ต้ อ ง",timeout=60000)
        assert error_message.is_visible()
        screenshot = capture_screenshot(page, test_name)
        save_result(test_name, "Pass", screenshot)
    except Exception:
        screenshot = capture_screenshot(page, test_name)
        save_result(test_name, "Fail", screenshot)
        raise

