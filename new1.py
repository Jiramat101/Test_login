import pytest
from playwright.sync_api import sync_playwright, Page
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image  # เพิ่มการนำเข้า Image
import os
import time

def setup_excel():
    filename = "TestSheet.xlsx"
    wb = load_workbook(filename)
    if "Test Results" not in wb.sheetnames:
        ws = wb.create_sheet("Test Results")
        ws.append(["Test Case", "Status", "Screenshot"])
        wb.save(filename)
    return filename

def save_result(test_name, status, screenshot_path):
    filename = setup_excel()
    wb = load_workbook(filename)
    ws = wb["Test Results"]
    
    
    
    # แทรกรูปภาพใน Excel
    if os.path.exists(screenshot_path):
        img = Image(screenshot_path)  # ใช้ Image จาก openpyxl.drawing.image
        img.width = 250  # ปรับขนาดรูปภาพ
        img.height = 190
        ws.add_image(img, f"I{new_row}")
    
    # บันทึกไฟล์ Excel
    wb.save(filename)

def capture_screenshot(page, test_name):
    screenshot_path = f"screenshots/{test_name}.png"
    
    # ตรวจสอบว่าโฟลเดอร์ screenshots/ มีอยู่หรือไม่ ถ้าไม่มีให้สร้าง
    os.makedirs("screenshots", exist_ok=True)
    
    try:
        # รอให้หน้าโหลดหรือองค์ประกอบที่ต้องการปรากฏก่อนจับภาพ
        page.wait_for_selector("input[type='submit'][value=' เข้าสู่ระบบ ']", timeout=10000)  # Adjust as necessary
        page.screenshot(path=screenshot_path)
        print(f"Screenshot saved to {screenshot_path}")  # ตรวจสอบว่าไฟล์ถูกบันทึกหรือไม่
    except Exception as e:
        print(f"Error capturing screenshot: {e}")
        screenshot_path = "Screenshot Failed"
    
    return screenshot_path

@pytest.fixture(scope="function")
def browser():
    with sync_playwright() as p:
        browser = p.chromium.launch(headless=False)
        yield browser
        browser.close()

@pytest.fixture(scope="function")
def page(browser):
    context = browser.new_context()
    page = context.new_page()
    yield page
    context.close()

def test_login01_pos01(page: Page):
    test_name = "test_login01_pos01"
    try:
        page.goto("https://reg.rmutk.ac.th/registrar/login.asp")

        # รอให้มีข้อความ "เปลี่ยนรหัสผ่าน" ปรากฏ
        success_message = page.wait_for_selector(':has-text("เปลี่ยนรหัสผ่าน")', timeout=60000)
        assert success_message.is_visible()

        screenshot = capture_screenshot(page, test_name)
        save_result_custom_row(test_name, "Pass", screenshot, 9)  # บันทึกที่แถว I9
    except Exception:
        screenshot = capture_screenshot(page, test_name)
        save_result_custom_row(test_name, "Fail", screenshot, 9)  # บันทึกที่แถว I9
        raise

def test_login01_neg01(page: Page):
    test_name = "test_login01_neg01"
    try:
        page.goto("https://reg.rmutk.ac.th/registrar/login.asp")
        page.locator('input[name="f_uid"]').fill("User_id")
        page.locator('input[name="f_pwd"]').fill("5555")
        page.locator('input[name="f_idcard"]').fill("5555")
        page.locator('input[type="submit"][value=" เข้าสู่ระบบ "]').click()

        error_message = page.wait_for_selector(
            "text=ก รุ ณ า ป้ อ น ร หั ส ป ร ะ จ ำ ตั ว น ศ .ร หั ส บั ต ร ป ร ะ ช า ช น แ ล ะ ร หั ส ผ่ า น ใ ห้ ถู ก ต้ อ ง",
            timeout=60000
        )
        color = error_message.evaluate('element => window.getComputedStyle(element).color')
        assert error_message.is_visible()
        assert color == "rgb(0,0,255)"

        screenshot = capture_screenshot(page, test_name)
        save_result_custom_row(test_name, "Pass", screenshot, 14)  # บันทึกที่แถว I14
    except Exception:
        screenshot = capture_screenshot(page, test_name)
        save_result_custom_row(test_name, "Fail", screenshot, 14)  # บันทึกที่แถว I14
        raise


def test_login01_neg02(page: Page):
    test_name = "test_login01_neg02"
    try:
        page.goto("https://reg.rmutk.ac.th/registrar/login.asp")
        page.locator('input[name="f_uid"]').fill("")
        page.locator('input[name="f_pwd"]').fill("")
        page.locator('input[name="f_idcard"]').fill("")  
        page.locator('input[type="submit"][value=" เข้าสู่ระบบ "]').click()

        # ตรวจสอบว่ามีข้อความแจ้งเตือน
        error_message = page.wait_for_selector("text=กรุณาป้อนรหัสประจำตัวและรหัสผ่านให้ถูกต้อง",timeout=60000)
        assert error_message.is_visible()

        screenshot = capture_screenshot(page, test_name)
        save_result_custom_row(test_name, "Pass", screenshot, 15)  # บันทึกที่แถว I15
    except Exception:
        screenshot = capture_screenshot(page, test_name)
        save_result_custom_row(test_name, "Fail", screenshot, 15)  # บันทึกที่แถว I15
        raise

def test_login01_neg03(page: Page):
    test_name = "test_login01_neg03"
    try:
        page.goto("https://reg.rmutk.ac.th/registrar/login.asp")
        page.locator('input[name="f_uid"]').fill("Usrt_id")
        page.locator('input[name="f_pwd"]').fill("รหัสผ่านจริง")
        page.locator('input[name="f_idcard"]').fill("5555")  # ใส่เลขบัตรประชาชนผิด
        page.locator('input[type="submit"][value=" เข้าสู่ระบบ "]').click()

        error_message = page.wait_for_selector("text=ก รุ ณ า ป้ อ น ร หั ส ป ร ะ จ ำ ตั ว น ศ .ร หั ส บั ต ร ป ร ะ ช า ช น แ ล ะ ร หั ส ผ่ า น ใ ห้ ถู ก ต้ อ ง",timeout=60000)
        assert error_message.is_visible()

        screenshot = capture_screenshot(page, test_name)
        save_result_custom_row(test_name, "Pass", screenshot, 16)  # บันทึกที่แถว I16
    except Exception:
        screenshot = capture_screenshot(page, test_name)
        save_result_custom_row(test_name, "Fail", screenshot, 16)  # บันทึกที่แถว I16
        raise

def test_login01_neg04(page: Page):
    test_name = "test_login01_neg04"
    try:
        page.goto("https://reg.rmutk.ac.th/registrar/login.asp")
        page.locator('input[name="f_uid"]').fill("5555")  # ใส่รหัสนักศึกษาผิด
        page.locator('input[name="f_pwd"]').fill("รัสผ่านจริง")  # ใส่รหัสผ่านถูก
        page.locator('input[name="f_idcard"]').fill("รหัส บปช")  # ใส่เลขบัตรประชาชนถูก
        page.locator('input[type="submit"][value=" เข้าสู่ระบบ "]').click()

        error_message = page.wait_for_selector("text=กรุณาป้อนรหัสประจำตัวและรหัสผ่านให้ถูกต้อง",timeout=60000)
        assert error_message.is_visible()

        screenshot = capture_screenshot(page, test_name)
        save_result_custom_row(test_name, "Pass", screenshot, 17)  # บันทึกที่แถว I17
    except Exception:
        screenshot = capture_screenshot(page, test_name)
        save_result_custom_row(test_name, "Fail", screenshot, 17)  # บันทึกที่แถว I17
        raise

def save_result_custom_row(test_name, status, screenshot_path, row_number):
    """ บันทึกผลลัพธ์ลง Excel ที่ตำแหน่งแถวที่กำหนด """
    filename = setup_excel()
    wb = load_workbook(filename)
    ws = wb["Test Results"]

    # บันทึกข้อมูลที่ column H (status) และ column I (screenshot) ในแถวที่กำหนด
    ws[f"H{row_number}"] = status

    # แสดงข้อมูลในคอนโซลเพื่อการดีบั๊ก
    print(f"Status: {status}")
    print(f"Screenshot Path: {screenshot_path}")

    # แทรกรูปภาพใน Excel
    if os.path.exists(screenshot_path):
        img = Image(screenshot_path)
        img.width = 250
        img.height = 190
        ws.add_image(img, f"I{row_number}")

    # บันทึกไฟล์ Excel

    wb.save(filename)
