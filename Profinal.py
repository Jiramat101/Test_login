import pytest
from playwright.sync_api import sync_playwright, Page
from openpyxl import load_workbook
import os

# กำหนดชื่อไฟล์ Excel
EXCEL_FILE = "C:\\Users\\VICTUS\\Documents\\งาน\\Scripting and Automation Tools\\TestSheet.xlsx"
SCREENSHOT_DIR = "screenshots"

# สร้างโฟลเดอร์สำหรับเก็บภาพหน้าจอหากยังไม่มี
os.makedirs(SCREENSHOT_DIR, exist_ok=True)

# ฟังก์ชันอัปเดตสถานะลงใน Excel
def update_test_status(test_name, status, screenshot_path):
    try:
        wb = load_workbook(EXCEL_FILE)
        ws = wb.active

        # ค้นหาแถวของ test_name
        for row in ws.iter_rows(min_row=2, values_only=False):  # ข้าม header
            if row[0].value == test_name:  # สมมติว่า test_name อยู่ที่คอลัมน์ A
                row[1].value = status  # คอลัมน์ B สำหรับสถานะ
                row[2].value = screenshot_path  # คอลัมน์ C สำหรับลิงก์ภาพ
                break
        
        wb.save(EXCEL_FILE)
        wb.close()
    except Exception as e:
        print(f"Error updating Excel: {e}")

# Fixture สำหรับเปิด Browser
@pytest.fixture(scope="function")
def page():
    with sync_playwright() as p:
        browser = p.chromium.launch(headless=False)
        context = browser.new_context()
        page = context.new_page()
        yield page
        context.close()
        browser.close()

# Test Case - Login ผิดพลาด
def test_login_fail(page: Page):
    test_name = "test_login_fail"
    page.goto("https://reg.rmutk.ac.th/registrar/login.asp")
    page.locator('input[name="f_uid"]').fill("65502100033-7")
    page.locator('input[name="f_pwd"]').fill("5555")
    page.locator('input[name="f_idcard"]').fill("5555")
    page.locator('input[type="submit"][value=" เข้าสู่ระบบ "]').click()
    
    try:
        error_message = page.get_by_text("กรุณาป้อนรหัสประจำตัวและรหัสผ่านให้ถูกต้อง")
        assert error_message.is_visible()

        # ถ้าผ่าน ให้บันทึกสถานะเป็น Pass
        status = "Pass"
    except AssertionError:
        # ถ้าผิดพลาด ให้บันทึกสถานะเป็น Fail
        status = "Fail"

    # บันทึก Screenshot
    screenshot_path = f"{SCREENSHOT_DIR}/{test_name}.png"
    page.screenshot(path=screenshot_path)

    # อัปเดตสถานะลงใน Excel
    update_test_status(test_name, status, screenshot_path)
'''
# Test Case - Login สำเร็จ
def test_login_success(page: Page):
    test_name = "test_login_success"
    page.goto("https://reg.rmutk.ac.th/registrar/login.asp")
    page.locator('input[name="f_uid"]').fill("ValidUser")
    page.locator('input[name="f_pwd"]').fill("ValidPass")
    page.locator('input[type="submit"][value=" เข้าสู่ระบบ "]').click()

    try:
        page.wait_for_url("**/main.asp**", timeout=5000)
        assert "main.asp" in page.url

        # ถ้าผ่าน ให้บันทึกสถานะเป็น Pass
        status = "Pass"
    except AssertionError:
        # ถ้าผิดพลาด ให้บันทึกสถานะเป็น Fail
        status = "Fail"

    # บันทึก Screenshot
    screenshot_path = f"{SCREENSHOT_DIR}/{test_name}.png"
    page.screenshot(path=screenshot_path)

    # อัปเดตสถานะลงใน Excel
    update_test_status(test_name, status, screenshot_path)
'''